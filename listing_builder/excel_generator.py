# /Users/shawn/Projects/Amazon/amazon-master-tool/listing_builder/excel_generator.py
# Purpose: Generate professional Excel reports with analysis and conditional formatting
# NOT for: Basic CSV exports - this is for rich formatted Excel with charts and colors

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference, PieChart
from datetime import datetime


def create_excel_report(analysis_data: dict, output_path: str = "product_analysis.xlsx"):
    """
    Create professional Excel report with multiple sheets and analysis.
    WHY: Provides comprehensive visual analysis of product opportunities.
    """
    # WHY: Create Excel writer
    writer = pd.ExcelWriter(output_path, engine='openpyxl')

    # WHY: Extract data
    all_products = analysis_data['all_products']
    top_product = analysis_data['top_product']
    num_competitors = analysis_data['num_competitors']
    avg_revenue = analysis_data['avg_revenue']
    total_revenue = analysis_data['total_niche_revenue']
    avg_rating = analysis_data['avg_rating']

    # WHY: Sort products by revenue
    sorted_products = sorted(all_products, key=lambda x: x['monthly_revenue'], reverse=True)

    # ========== SHEET 1: SUMMARY ==========
    summary_data = {
        'Metric': [
            'Total Products',
            'Total Niche Revenue/Month',
            'Average Revenue/Product',
            'Average Rating',
            'Top Seller ASIN',
            'Top Seller Revenue',
            'Top Seller Reviews',
            'Top Seller Rating',
            'Top Seller BSR',
            'Analysis Date'
        ],
        'Value': [
            num_competitors,
            f"€{total_revenue:,.2f}",
            f"€{avg_revenue:,.2f}",
            f"{avg_rating:.2f} ⭐",
            top_product.get('asin', 'N/A'),
            f"€{top_product.get('monthly_revenue', 0):,.2f}",
            f"{top_product.get('review_count', 0):,}",
            f"{top_product.get('review_rating', 0):.1f} ⭐",
            f"#{top_product.get('bsr_rank', 999999):,}",
            datetime.now().strftime('%Y-%m-%d %H:%M')
        ]
    }
    summary_df = pd.DataFrame(summary_data)
    summary_df.to_excel(writer, sheet_name='Summary', index=False)

    # ========== SHEET 2: ALL PRODUCTS ==========
    products_data = []
    for i, prod in enumerate(sorted_products, 1):
        products_data.append({
            'Rank': i,
            'ASIN': prod['asin'],
            'Title': prod['title'][:80] + '...' if len(prod['title']) > 80 else prod['title'],
            'Monthly Revenue (€)': prod['monthly_revenue'],
            'Reviews': prod['review_count'],
            'Rating': prod['review_rating'],
            'BSR': prod['bsr_rank'],
            'Price (€)': prod['price'],
            'Category': prod.get('category', 'N/A'),
            'Revenue Score': _calculate_revenue_score(prod['monthly_revenue']),
            'Competition Score': _calculate_competition_score(prod['review_count']),
            'Quality Score': _calculate_quality_score(prod['review_rating']),
            'Demand Score': _calculate_demand_score(prod['bsr_rank']),
            'Overall Score': _calculate_overall_score(prod)
        })

    products_df = pd.DataFrame(products_data)
    products_df.to_excel(writer, sheet_name='All Products', index=False)

    # ========== SHEET 3: TOP 10 ANALYSIS ==========
    top10_data = []
    for i, prod in enumerate(sorted_products[:10], 1):
        # WHY: Detailed analysis for top 10
        revenue_per_review = prod['monthly_revenue'] / max(prod['review_count'], 1)

        analysis = {
            'Rank': i,
            'ASIN': prod['asin'],
            'Title': prod['title'][:100],
            'Monthly Revenue (€)': prod['monthly_revenue'],
            'Reviews': prod['review_count'],
            'Rating': prod['review_rating'],
            'BSR': prod['bsr_rank'],
            'Price (€)': prod['price'],
            'Revenue/Review': revenue_per_review,
            'Est. Units/Month': int(prod['monthly_revenue'] / max(prod['price'], 1)),
            'Market Share (%)': (prod['monthly_revenue'] / total_revenue * 100) if total_revenue > 0 else 0,
            'Competition Level': _get_competition_level(prod['review_count']),
            'Entry Difficulty': _get_entry_difficulty(prod),
            'Profit Potential': _get_profit_potential(prod),
            'Recommendation': _get_recommendation(prod, num_competitors)
        }
        top10_data.append(analysis)

    top10_df = pd.DataFrame(top10_data)
    top10_df.to_excel(writer, sheet_name='Top 10 Analysis', index=False)

    # ========== SHEET 4: RECOMMENDATIONS ==========
    recommendations = _generate_recommendations(sorted_products, num_competitors, total_revenue)
    rec_df = pd.DataFrame(recommendations)
    rec_df.to_excel(writer, sheet_name='Recommendations', index=False)

    # WHY: Save and get workbook for formatting
    writer.close()

    # WHY: Apply formatting
    _apply_formatting(output_path)

    return output_path


def _calculate_revenue_score(revenue: float) -> int:
    """Calculate revenue score 0-100."""
    if revenue >= 50000: return 100
    elif revenue >= 20000: return 85
    elif revenue >= 10000: return 70
    elif revenue >= 5000: return 50
    elif revenue >= 1000: return 30
    else: return 10


def _calculate_competition_score(reviews: int) -> int:
    """Calculate competition score 0-100 (lower reviews = higher score)."""
    if reviews < 50: return 100
    elif reviews < 200: return 80
    elif reviews < 500: return 60
    elif reviews < 1000: return 40
    else: return 20


def _calculate_quality_score(rating: float) -> int:
    """Calculate quality score 0-100."""
    if rating >= 4.5: return 100
    elif rating >= 4.0: return 80
    elif rating >= 3.5: return 60
    elif rating >= 3.0: return 40
    else: return 20


def _calculate_demand_score(bsr: int) -> int:
    """Calculate demand score 0-100."""
    if bsr < 5000: return 100
    elif bsr < 20000: return 80
    elif bsr < 50000: return 60
    elif bsr < 100000: return 40
    else: return 20


def _calculate_overall_score(product: dict) -> int:
    """Calculate weighted overall score."""
    revenue_score = _calculate_revenue_score(product['monthly_revenue'])
    competition_score = _calculate_competition_score(product['review_count'])
    quality_score = _calculate_quality_score(product['review_rating'])
    demand_score = _calculate_demand_score(product['bsr_rank'])

    # WHY: Weighted average (revenue 30%, competition 30%, quality 20%, demand 20%)
    overall = (revenue_score * 0.3 + competition_score * 0.3 +
               quality_score * 0.2 + demand_score * 0.2)
    return int(overall)


def _get_competition_level(reviews: int) -> str:
    """Get competition level text."""
    if reviews < 50: return "VERY LOW"
    elif reviews < 200: return "LOW"
    elif reviews < 500: return "MEDIUM"
    elif reviews < 1000: return "HIGH"
    else: return "VERY HIGH"


def _get_entry_difficulty(product: dict) -> str:
    """Assess entry difficulty."""
    reviews = product['review_count']
    revenue = product['monthly_revenue']

    if reviews > 1000 and revenue > 10000:
        return "VERY HARD"
    elif reviews > 500 or revenue > 20000:
        return "HARD"
    elif reviews > 200 or revenue > 5000:
        return "MEDIUM"
    else:
        return "EASY"


def _get_profit_potential(product: dict) -> str:
    """Assess profit potential."""
    revenue = product['monthly_revenue']
    reviews = product['review_count']
    rating = product['review_rating']

    score = _calculate_overall_score(product)

    if score >= 75:
        return "HIGH"
    elif score >= 50:
        return "MEDIUM"
    else:
        return "LOW"


def _get_recommendation(product: dict, total_competitors: int) -> str:
    """Get go/no-go recommendation."""
    score = _calculate_overall_score(product)
    reviews = product['review_count']
    revenue = product['monthly_revenue']

    if score >= 70 and reviews < 200 and revenue > 5000:
        return "✅ GO"
    elif score >= 50 and reviews < 500:
        return "⚠️ MAYBE"
    else:
        return "❌ NO GO"


def _generate_recommendations(products: list, num_competitors: int, total_revenue: float) -> list:
    """Generate strategic recommendations."""
    recommendations = []

    # WHY: Niche overview
    recommendations.append({
        'Category': 'NICHE OVERVIEW',
        'Insight': f'{num_competitors} competitors, €{total_revenue:,.0f}/month total revenue',
        'Recommendation': 'Analyze top 10 products for opportunities',
        'Priority': 'HIGH'
    })

    # WHY: Find low-hanging fruit
    easy_targets = [p for p in products if p['review_count'] < 50 and p['monthly_revenue'] > 1000]
    if easy_targets:
        recommendations.append({
            'Category': 'LOW COMPETITION',
            'Insight': f'Found {len(easy_targets)} products with <50 reviews',
            'Recommendation': f'Target: {easy_targets[0]["asin"]} (€{easy_targets[0]["monthly_revenue"]:,.0f}/mo)',
            'Priority': 'HIGH'
        })

    # WHY: Quality gaps
    poor_rating_products = [p for p in products[:20] if p['review_rating'] < 4.0 and p['monthly_revenue'] > 2000]
    if poor_rating_products:
        recommendations.append({
            'Category': 'QUALITY OPPORTUNITY',
            'Insight': f'{len(poor_rating_products)} high-revenue products with <4.0 rating',
            'Recommendation': 'Improve quality and capture dissatisfied customers',
            'Priority': 'MEDIUM'
        })

    # WHY: Market concentration
    top5_revenue = sum(p['monthly_revenue'] for p in products[:5])
    concentration = (top5_revenue / total_revenue * 100) if total_revenue > 0 else 0

    if concentration > 70:
        recommendations.append({
            'Category': 'MARKET CONCENTRATION',
            'Insight': f'Top 5 products = {concentration:.0f}% of market',
            'Recommendation': 'Dominated niche - target long-tail keywords',
            'Priority': 'MEDIUM'
        })
    else:
        recommendations.append({
            'Category': 'MARKET FRAGMENTATION',
            'Insight': f'Top 5 = {concentration:.0f}% (fragmented market)',
            'Recommendation': 'Multiple opportunities available',
            'Priority': 'HIGH'
        })

    # WHY: Price analysis
    avg_price = sum(p['price'] for p in products[:20]) / 20 if products else 0
    recommendations.append({
        'Category': 'PRICING STRATEGY',
        'Insight': f'Average price: €{avg_price:.2f}',
        'Recommendation': f'Price competitively: €{avg_price * 0.9:.2f}-€{avg_price * 1.1:.2f}',
        'Priority': 'MEDIUM'
    })

    return recommendations


def _apply_formatting(file_path: str):
    """Apply Excel formatting with colors and styles."""
    from openpyxl import load_workbook

    wb = load_workbook(file_path)

    # WHY: Define colors
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    high_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    medium_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    low_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

    header_font = Font(bold=True, color='FFFFFF', size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # WHY: Format Summary sheet
    if 'Summary' in wb.sheetnames:
        ws = wb['Summary']
        for row in ws['A1:B1']:
            for cell in row:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 30

    # WHY: Format All Products sheet
    if 'All Products' in wb.sheetnames:
        ws = wb['All Products']

        # Header
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Conditional formatting for Overall Score
        for row in range(2, ws.max_row + 1):
            score_cell = ws[f'N{row}']  # Overall Score column
            if score_cell.value:
                try:
                    score = float(score_cell.value)
                    if score >= 70:
                        score_cell.fill = high_fill
                    elif score >= 50:
                        score_cell.fill = medium_fill
                    else:
                        score_cell.fill = low_fill
                except:
                    pass

        # Auto-width
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    # WHY: Format Top 10 Analysis
    if 'Top 10 Analysis' in wb.sheetnames:
        ws = wb['Top 10 Analysis']

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Color recommendations
        for row in range(2, ws.max_row + 1):
            rec_cell = ws[f'O{row}']  # Recommendation column
            if rec_cell.value:
                if '✅' in str(rec_cell.value):
                    rec_cell.fill = high_fill
                elif '⚠️' in str(rec_cell.value):
                    rec_cell.fill = medium_fill
                elif '❌' in str(rec_cell.value):
                    rec_cell.fill = low_fill

    # WHY: Format Recommendations
    if 'Recommendations' in wb.sheetnames:
        ws = wb['Recommendations']

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        for row in range(2, ws.max_row + 1):
            priority_cell = ws[f'D{row}']
            if priority_cell.value == 'HIGH':
                for col in ['A', 'B', 'C', 'D']:
                    ws[f'{col}{row}'].fill = high_fill
            elif priority_cell.value == 'MEDIUM':
                for col in ['A', 'B', 'C', 'D']:
                    ws[f'{col}{row}'].fill = medium_fill

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 15

    wb.save(file_path)
