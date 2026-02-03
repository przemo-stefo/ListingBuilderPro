# /Users/shawn/Projects/ListingBuilderPro/listing_builder/backend/services/compliance/parser_amazon.py
# Purpose: Parse Amazon Flat File XLSM template (Row 3 = field IDs, 300+ cols)
# NOT for: Validation rules or report generation

from typing import List, Dict
import structlog

logger = structlog.get_logger()


def parse_amazon_xlsm(file_bytes: bytes) -> List[Dict[str, str]]:
    """
    Parse Amazon Flat File XLSM (Seller Central template with macros).

    Format:
      Row 1 = human-readable labels
      Row 2 = blank or instructions
      Row 3 = machine-readable field IDs (our column keys)
      Row 4+ = product data

    Uses Row 3 field IDs as dict keys so column positions don't matter.
    Handles XLSM (macro-enabled) via openpyxl's keep_vba=False.

    Returns list of dicts, one per product row.
    Keys are the Row 3 field IDs (e.g. 'item_sku', 'manufacturer', 'battery_type').
    """
    import openpyxl

    from io import BytesIO
    wb = openpyxl.load_workbook(
        BytesIO(file_bytes),
        read_only=True,
        data_only=True,
        # WHY keep_vba not set: read_only mode ignores macros automatically
    )

    # Amazon templates usually have a "Template" sheet, fall back to active
    ws = None
    for name in wb.sheetnames:
        if "template" in name.lower():
            ws = wb[name]
            break
    if ws is None:
        ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 4:
        logger.warning("amazon_xlsm_too_short", row_count=len(rows))
        return []

    # Row 3 (index 2) = field IDs — these are the canonical column names
    raw_field_ids = rows[2]
    field_ids = [str(fid).strip() if fid is not None else "" for fid in raw_field_ids]

    # Build column index → field_id mapping (skip empty field IDs)
    col_map = {}
    for col_idx, fid in enumerate(field_ids):
        if fid:
            col_map[col_idx] = fid

    products = []
    for row_idx, row in enumerate(rows[3:], start=4):
        # Skip empty rows — check if any mapped column has data
        has_data = False
        for ci in col_map:
            if ci < len(row) and row[ci] is not None and str(row[ci]).strip():
                has_data = True
                break
        if not has_data:
            continue

        product = {"_row_number": row_idx}
        for col_idx, fid in col_map.items():
            val = row[col_idx] if col_idx < len(row) else None
            product[fid] = str(val).strip() if val is not None else ""

        products.append(product)

    logger.info("amazon_xlsm_parsed", products=len(products), fields=len(col_map))
    return products
