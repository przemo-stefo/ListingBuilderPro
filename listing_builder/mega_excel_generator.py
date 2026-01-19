# /Users/shawn/Projects/Amazon/amazon-master-tool/listing_builder/mega_excel_generator.py
# Purpose: Generate comprehensive Excel comparison report from multiple data sources
# NOT for: Single dataset analysis - use excel_generator.py instead

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime


def create_comparison_excel(comparison_data: dict, output_path: str = "isolbau_mega_analysis.xlsx"):
    """
    Create comprehensive Excel comparing Niche CSV + Black Box data.
    WHY: Shows complete market picture from multiple data sources.
    """
    writer = pd.ExcelWriter(output_path, engine='openpyxl')

    # ========== SHEET 1: EXECUTIVE SUMMARY ==========
    summary_data = {
        'Metric': [
            'Total Unique Products',
            'Niche CSV Products',
            'Black Box Products',
            'Products in Both (Overlap)',
            'Niche-Only Products',
            'Black Box-Only Products',
            '',
            'Combined Total Revenue/Month',
            'Niche CSV Revenue',
            'Black Box Revenue',
            '',
            'Analysis Date'
        ],
        'Value': [
            comparison_data['total_unique_products'],
            comparison_data['niche_count'],
            comparison_data['blackbox_count'],
            comparison_data['overlap_count'],
            comparison_data['niche_only_count'],
            comparison_data['blackbox_only_count'],
            '',
            f"€{comparison_data['combined_total_revenue']:,.0f}",
            f"€{sum(p['monthly_revenue'] for p in comparison_data['niche_only']) + sum(o['niche_data']['monthly_revenue'] for o in comparison_data['overlap']):,.0f}",
            f"€{sum(p['monthly_revenue'] for p in comparison_data['blackbox_only']) + sum(o['blackbox_data']['monthly_revenue'] for o in comparison_data['overlap']):,.0f}",
            '',
            datetime.now().strftime('%Y-%m-%d %H:%M')
        ]
    }
    summary_df = pd.DataFrame(summary_data)
    summary_df.to_excel(writer, sheet_name='Executive Summary', index=False)

    # ========== SHEET 2: ALL 300 UNIQUE PRODUCTS ==========
    all_products_data = []
    for i, prod in enumerate(sorted(comparison_data['all_unique_products'],
                                    key=lambda x: x['monthly_revenue'], reverse=True), 1):
        all_products_data.append({
            'Rank': i,
            'ASIN': prod['asin'],
            'Title': prod['title'][:80] + '...' if len(prod['title']) > 80 else prod['title'],
            'Monthly Revenue (€)': prod['monthly_revenue'],
            'Reviews': prod['review_count'],
            'Rating': prod['review_rating'],
            'BSR': prod['bsr_rank'],
            'Price (€)': prod['price'],
            'Category': prod.get('category', 'N/A')
        })

    all_df = pd.DataFrame(all_products_data)
    all_df.to_excel(writer, sheet_name='All 300 Products', index=False)

    # ========== SHEET 3: DATA SOURCE COMPARISON (OVERLAP) ==========
    if comparison_data['overlap']:
        overlap_data = []
        for i, item in enumerate(comparison_data['overlap'], 1):
            nd = item['niche_data']
            bd = item['blackbox_data']
            rev_diff = nd['monthly_revenue'] - bd['monthly_revenue']
            rev_diff_pct = (rev_diff / max(bd['monthly_revenue'], 1)) * 100

            overlap_data.append({
                'Rank': i,
                'ASIN': item['asin'],
                'Title': nd['title'][:60],
                'Niche Revenue (€)': nd['monthly_revenue'],
                'Black Box Revenue (€)': bd['monthly_revenue'],
                'Difference (€)': rev_diff,
                'Difference (%)': rev_diff_pct,
                'Niche Reviews': nd['review_count'],
                'Black Box Reviews': bd['review_count'],
                'Niche Rating': nd['review_rating'],
                'Black Box Rating': bd['review_rating'],
                'Niche BSR': nd['bsr_rank'],
                'Black Box BSR': bd['bsr_rank']
            })

        overlap_df = pd.DataFrame(overlap_data)
        overlap_df.to_excel(writer, sheet_name='Data Comparison (15)', index=False)

    # ========== SHEET 4: NICHE-ONLY PRODUCTS (80) ==========
    niche_only_data = []
    for i, prod in enumerate(sorted(comparison_data['niche_only'],
                                   key=lambda x: x['monthly_revenue'], reverse=True), 1):
        niche_only_data.append({
            'Rank': i,
            'ASIN': prod['asin'],
            'Title': prod['title'][:80],
            'Monthly Revenue (€)': prod['monthly_revenue'],
            'Reviews': prod['review_count'],
            'Rating': prod['review_rating'],
            'BSR': prod['bsr_rank'],
            'Price (€)': prod['price']
        })

    niche_df = pd.DataFrame(niche_only_data)
    niche_df.to_excel(writer, sheet_name='Niche-Only (80)', index=False)

    # ========== SHEET 5: BLACK BOX-ONLY PRODUCTS (205) ==========
    bb_only_data = []
    for i, prod in enumerate(sorted(comparison_data['blackbox_only'],
                                   key=lambda x: x['monthly_revenue'], reverse=True), 1):
        bb_only_data.append({
            'Rank': i,
            'ASIN': prod['asin'],
            'Title': prod['title'][:80],
            'Monthly Revenue (€)': prod['monthly_revenue'],
            'Reviews': prod['review_count'],
            'Rating': prod['review_rating'],
            'BSR': prod['bsr_rank'],
            'Price (€)': prod['price']
        })

    bb_df = pd.DataFrame(bb_only_data)
    bb_df.to_excel(writer, sheet_name='Black Box-Only (205)', index=False)

    # ========== SHEET 6: TOP 50 BY REVENUE ==========
    top50_data = []
    sorted_all = sorted(comparison_data['all_unique_products'],
                       key=lambda x: x['monthly_revenue'], reverse=True)[:50]

    for i, prod in enumerate(sorted_all, 1):
        # WHY: Determine data source
        source = 'Both'
        if prod['asin'] in [p['asin'] for p in comparison_data['niche_only']]:
            source = 'Niche Only'
        elif prod['asin'] in [p['asin'] for p in comparison_data['blackbox_only']]:
            source = 'Black Box Only'

        top50_data.append({
            'Rank': i,
            'ASIN': prod['asin'],
            'Title': prod['title'][:70],
            'Monthly Revenue (€)': prod['monthly_revenue'],
            'Reviews': prod['review_count'],
            'Rating': prod['review_rating'],
            'BSR': prod['bsr_rank'],
            'Price (€)': prod['price'],
            'Data Source': source,
            'Market Share (%)': (prod['monthly_revenue'] / comparison_data['combined_total_revenue'] * 100)
        })

    top50_df = pd.DataFrame(top50_data)
    top50_df.to_excel(writer, sheet_name='Top 50 by Revenue', index=False)

    # ========== SHEET 7: STRATEGIC INSIGHTS ==========
    insights_data = []

    # WHY: Market concentration
    top10_revenue = sum(p['monthly_revenue'] for p in sorted_all[:10])
    concentration = (top10_revenue / comparison_data['combined_total_revenue'] * 100)

    insights_data.append({
        'Category': 'MARKET CONCENTRATION',
        'Insight': f'Top 10 products = {concentration:.1f}% of total revenue (€{top10_revenue:,.0f}/mo)',
        'Recommendation': 'Highly concentrated - focus on top performers or find long-tail opportunities',
        'Priority': 'HIGH' if concentration > 50 else 'MEDIUM'
    })

    # WHY: Data source reliability
    overlap_pct = (comparison_data['overlap_count'] / min(comparison_data['niche_count'], comparison_data['blackbox_count']) * 100)

    insights_data.append({
        'Category': 'DATA SOURCE OVERLAP',
        'Insight': f'Only {overlap_pct:.1f}% overlap between sources ({comparison_data["overlap_count"]} products)',
        'Recommendation': 'Low overlap suggests different search criteria - combine both for complete picture',
        'Priority': 'HIGH'
    })

    # WHY: Revenue opportunity
    avg_revenue_niche = sum(p['monthly_revenue'] for p in comparison_data['niche_only']) / max(len(comparison_data['niche_only']), 1)
    avg_revenue_bb = sum(p['monthly_revenue'] for p in comparison_data['blackbox_only']) / max(len(comparison_data['blackbox_only']), 1)

    insights_data.append({
        'Category': 'REVENUE PATTERNS',
        'Insight': f'Niche CSV avg: €{avg_revenue_niche:,.0f}/mo vs Black Box avg: €{avg_revenue_bb:,.0f}/mo',
        'Recommendation': f'Niche targets high-revenue products, Black Box shows wider market including long-tail',
        'Priority': 'MEDIUM'
    })

    # WHY: Entry opportunities
    low_competition = [p for p in comparison_data['all_unique_products']
                      if p['review_count'] < 100 and p['monthly_revenue'] > 1000]

    if low_competition:
        insights_data.append({
            'Category': 'LOW COMPETITION OPPORTUNITIES',
            'Insight': f'Found {len(low_competition)} products with <100 reviews and >€1,000/mo revenue',
            'Recommendation': f'Target: {low_competition[0]["asin"]} (€{low_competition[0]["monthly_revenue"]:,.0f}/mo, {low_competition[0]["review_count"]} reviews)',
            'Priority': 'HIGH'
        })

    insights_df = pd.DataFrame(insights_data)
    insights_df.to_excel(writer, sheet_name='Strategic Insights', index=False)

    writer.close()

    # WHY: Apply formatting
    _apply_mega_formatting(output_path)

    return output_path


def _apply_mega_formatting(file_path: str):
    """Apply professional Excel formatting."""
    wb = load_workbook(file_path)

    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)

    # WHY: Format all sheets
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Header row
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Auto-width
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 60)
            ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(file_path)
